import csv
import os
import re
import threading
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
try:
    import pandas as pd
except ImportError:
    pd = None
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    load_workbook = None
    PatternFill = Font = Alignment = None
EXPORT_COLUMNS = [
    "空",
    "公司名称(销售方)",
    "纳税人识别号(购买方税号)",
    "发票编码(发票号码)",
    "开票日期",
    "空2",
    "产品名称",
    "型号(规格型号)",
    "数量",
    "金额",
    "税额",
    "总价",
]
TEMPLATE_SPECS = {
    "cn_e_invoice_common_tall_v1": {
        "title_keywords": ["电子发票", "普通发票"],
        "height_min": 430,
        "regions": {
            "title": (180, 10, 420, 45),
            "meta_number": (430, 25, 585, 45),
            "meta_date": (430, 45, 585, 65),
            "buyer": (30, 90, 298, 147),
            "seller": (315, 90, 582, 147),
            "items_band": (12, 147, 583, 320),
            "grand_total": (12, 320, 583, 352),
            "note": (12, 352, 583, 408),
            "issuer": (40, 420, 140, 440),
        },
        "item_columns": {
            "item_name": (12, 114),
            "model": (114, 188),
            "unit": (188, 229),
            "quantity": (229, 297),
            "unit_price": (297, 392),
            "amount": (392, 446),
            "tax_rate": (446, 551),
            "tax_amount": (551, 583),
        },
        "anchor": {"x": (280, 291), "top": (155, 305), "row_top_offset": -0.8, "next_minus": 5.0, "last_bottom": 312},
    },
    "cn_vat_special_compact_v1": {
        "title_keywords": ["电子发票", "增值税专用发票"],
        "height_max": 430,
        "regions": {
            "title": (160, 18, 425, 45),
            "meta_number": (438, 28, 582, 43),
            "meta_date": (438, 45, 582, 60),
            "buyer": (32, 90, 298, 145),
            "seller": (317, 90, 582, 145),
            "items_band": (12.756, 147.406, 582.510, 272.128),
            "grand_total": (12.756, 272.128, 582.510, 294.805),
            "note": (12.756, 294.805, 582.510, 351.497),
            "issuer": (12, 360, 120, 392),
        },
        "item_columns": {
            "item_name": (12, 114),
            "model": (114, 188),
            "unit": (188, 229),
            "quantity": (229, 297),
            "unit_price": (297, 392),
            "amount": (392, 446),
            "tax_rate": (446, 551),
            "tax_amount": (551, 583),
        },
        "anchor": {"x": (281, 291), "top": (158, 230), "row_top_offset": -0.8, "next_minus": 3.0, "last_bottom": 258.5},
    },
    "cn_e_invoice_common_compact_v1": {
        "title_keywords": ["电子发票", "普通发票"],
        "height_max": 430,
        "regions": {
            "title": (160, 18, 425, 45),
            "meta_number": (438, 28, 582, 43),
            "meta_date": (438, 45, 582, 60),
            "buyer": (30, 90, 298, 145),
            "seller": (315, 90, 582, 145),
            "items_band": (12.756, 147.406, 582.510, 272.128),
            "grand_total": (12.756, 272.128, 582.510, 294.805),
            "note": (12.756, 294.805, 582.510, 351.497),
            "issuer": (12, 360, 120, 392),
        },
        "item_columns": {
            "item_name": (12, 114),
            "model": (114, 188),
            "unit": (188, 229),
            "quantity": (229, 293),
            "unit_price": (293, 392),
            "amount": (392, 446),
            "tax_rate": (446, 551),
            "tax_amount": (551, 583),
        },
        "anchor": {"x": (281, 291), "top": (158, 245), "row_top_offset": -0.8, "next_minus": 3.0, "last_bottom": 258.5},
    },
}
@dataclass
class InvoiceRow:
    source_file: str = ""
    invoice_number: str = ""
    invoice_date: str = ""
    buyer_name: str = ""
    buyer_tax_no: str = ""
    seller_name: str = ""
    seller_tax_no: str = ""
    item_name: str = ""
    model: str = ""
    unit: str = ""
    quantity: str = ""
    unit_price: str = ""
    amount: str = ""
    tax_rate: str = ""
    tax_amount: str = ""
    total: str = ""
    def to_export_dict(self) -> Dict[str, str]:
        return {
            "空": "",
            "公司名称(销售方)": self.seller_name,
            "纳税人识别号(购买方税号)": self.buyer_tax_no,
            "发票编码(发票号码)": self.invoice_number,
            "开票日期": self.invoice_date,
            "空2": "",
            "产品名称": self.item_name,
            "型号(规格型号)": self.model,
            "数量": self.quantity,
            "金额": self.amount,
            "税额": self.tax_amount,
            "总价": self.total,
        }
def normalize_text(text: str) -> str:
    text = text.replace("\u3000", " ")
    text = text.replace("：", ":")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[ \t]+", " ", text)
    return text
def crop_text(page, bbox: Tuple[float, float, float, float]) -> str:
    text = page.crop(bbox).extract_text(x_tolerance=1, y_tolerance=3) or ""
    return normalize_text(text).strip()
def detect_invoice_template(page) -> str:
    title = crop_text(page, (150, 8, 440, 55))
    title_compact = compact_label(title)
    h = float(page.height)
    if "增值税专用发票" in title_compact:
        return "cn_vat_special_compact_v1"
    if "普通发票" in title_compact:
        if h > 430:
            return "cn_e_invoice_common_tall_v1"
        return "cn_e_invoice_common_compact_v1"
    return ""
def extract_label_value(block_text: str, labels: List[str]) -> str:
    compact = compact_label(block_text)
    for label in labels:
        m = re.search(rf"{re.escape(label)}[:：]?(.+?)(?=(?:名称|统一社会信用代码/?纳税人识别号|纳税人识别号|地址|开户行|电话|账号|开票人|$))", compact)
        if m:
            return m.group(1).strip()
    return ""
def extract_anchors_from_quantity(page, anchor_spec: Dict[str, Any]) -> List[float]:
    words = page.extract_words(use_text_flow=True) or []
    x0, x1 = anchor_spec["x"]
    top0, top1 = anchor_spec["top"]
    anchors: List[float] = []
    for w in words:
        wx0 = float(w.get("x0", 0))
        wtop = float(w.get("top", 0))
        txt = (w.get("text") or "").strip()
        if x0 <= wx0 <= x1 and top0 <= wtop <= top1 and re.fullmatch(r"\d+(?:\.\d+)?", txt):
            anchors.append(round(wtop, 2))
    if not anchors:
        return []
    anchors = sorted(set(anchors))
    deduped = [anchors[0]]
    for val in anchors[1:]:
        if abs(val - deduped[-1]) > 1.5:
            deduped.append(val)
    return deduped
def extract_items_by_layout(page, spec: Dict[str, Any]) -> List[Dict[str, str]]:
    anchors = extract_anchors_from_quantity(page, spec["anchor"])
    if not anchors:
        return []
    item_columns = spec["item_columns"]
    row_top_offset = float(spec["anchor"]["row_top_offset"])
    next_minus = float(spec["anchor"]["next_minus"])
    last_bottom = float(spec["anchor"]["last_bottom"])
    rows: List[Dict[str, str]] = []
    for idx, anchor_top in enumerate(anchors):
        row_top = anchor_top + row_top_offset
        row_bottom = (anchors[idx + 1] - next_minus) if idx < len(anchors) - 1 else last_bottom
        record: Dict[str, str] = {}
        for field, (x0, x1) in item_columns.items():
            record[field] = crop_text(page, (x0, row_top, x1, row_bottom))
        record["item_name"] = record.get("item_name", "").replace("\n", "").strip().lstrip("*")
        rows.append({
            "item_name": record.get("item_name", ""),
            "model": record.get("model", "").strip(),
            "unit": record.get("unit", "").strip(),
            "quantity": first_match(r"(\d+(?:\.\d+)?)", record.get("quantity", ""), group=1),
            "unit_price": first_match(r"([0-9]+(?:\.[0-9]+)?)", clean_money(record.get("unit_price", "")), group=1),
            "amount": first_match(r"([0-9]+(?:\.[0-9]+)?)", clean_money(record.get("amount", "")), group=1),
            "tax_rate": first_match(r"([0-9]+(?:\.[0-9]+)?%?)", record.get("tax_rate", ""), group=1),
            "tax_amount": first_match(r"([0-9]+(?:\.[0-9]+)?)", clean_money(record.get("tax_amount", "")), group=1),
        })
    cleaned_rows = [r for r in rows if any(r.get(k) for k in ["item_name", "amount", "tax_amount", "quantity"])]
    return cleaned_rows
def extract_structured_invoice(pdf_path: str) -> Optional[Dict[str, Any]]:
    if pdfplumber is None:
        return None
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return None
        page = pdf.pages[0]
        template_id = detect_invoice_template(page)
        if not template_id:
            return None
        spec = TEMPLATE_SPECS.get(template_id)
        if not spec:
            return None
        regions = spec["regions"]
        buyer_text = crop_text(page, regions["buyer"])
        seller_text = crop_text(page, regions["seller"])
        grand_total_text = crop_text(page, regions["grand_total"])
        items_band_text = crop_text(page, regions["items_band"])
        meta_number_text = crop_text(page, regions["meta_number"])
        meta_date_text = crop_text(page, regions["meta_date"])
        invoice_number = first_match(r"发票号码[:：]?\s*([0-9A-Za-z]+)", meta_number_text)
        invoice_date = first_match(r"开票日期[:：]?\s*([0-9]{4}[-/年][0-9]{1,2}[-/月][0-9]{1,2}日?)", meta_date_text)
        buyer_name = first_match(r"名称[:：]?(.+?)(?:统一社会信用代码|纳税人识别号|$)", compact_label(buyer_text))
        seller_name = first_match(r"名称[:：]?(.+?)(?:统一社会信用代码|纳税人识别号|$)", compact_label(seller_text))
        buyer_tax_no = first_match(r"(?:统一社会信用代码/?纳税人识别号|纳税人识别号)[:：]?([0-9A-Z]{15,20})", compact_label(buyer_text))
        seller_tax_no = first_match(r"(?:统一社会信用代码/?纳税人识别号|纳税人识别号)[:：]?([0-9A-Z]{15,20})", compact_label(seller_text))
        money_vals = re.findall(r"([0-9]+\.[0-9]+)", f"{items_band_text}\n{grand_total_text}")
        total = ""
        if money_vals:
            total = money_vals[-1]
        items = extract_items_by_layout(page, spec)
        return {
            "header": {
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "buyer_name": buyer_name,
                "buyer_tax_no": buyer_tax_no,
                "seller_name": seller_name,
                "seller_tax_no": seller_tax_no,
                "total": total,
            },
            "items": items,
        }
def extract_text_from_pdf(pdf_path: str) -> str:
    if pdfplumber is None:
        raise RuntimeError("缺少依赖: pdfplumber，请先安装 (pip install pdfplumber)")
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            pages.append(t)
    return normalize_text("\n".join(pages))
def first_match(pattern: str, text: str, flags: int = 0, group: int = 1) -> str:
    """Return regex match content safely.
    Defaults to capturing group 1 for historical compatibility. If the pattern
    has no capturing groups, it automatically falls back to the full match.
    """
    m = re.search(pattern, text, flags)
    if not m:
        return ""
    if group == 0:
        return m.group(0).strip()
    if m.lastindex is None:
        return m.group(0).strip()
    if group <= m.lastindex:
        return m.group(group).strip()
    return ""
def compact_label(text: str) -> str:
    """Remove whitespace/newlines for matching vertical or split labels."""
    return re.sub(r"\s+", "", text)
def clean_money(value: str) -> str:
    value = value.replace("￥", "").replace("¥", "").strip()
    return value
def split_name_model_unit(item_name: str, model: str, unit: str) -> Dict[str, str]:
    """Split merged item text into item/model/unit when possible.
    仅在“紧凑无空格串行文本”中尝试推断，避免把正常产品描述误拆。
    """
    item_name = (item_name or "").strip()
    model = (model or "").strip()
    unit = (unit or "").strip()
    if not item_name:
        return {"item_name": item_name, "model": model, "unit": unit}
    compact_item = re.sub(r"\s+", "", item_name)
    has_spacing = bool(re.search(r"\s", item_name))
    def looks_like_unit_token(token: str) -> bool:
        token = (token or "").strip()
        if not token:
            return False
        if re.search(r"\d", token):
            return False
        if re.search(r"[*/_=+~`!@#$%^&?,;:'\"\\|]", token):
            return False
        if re.fullmatch(r"[A-Za-z]{1,4}", token):
            return True
        if re.fullmatch(r"[\u4e00-\u9fff]{1,3}", token):
            return True
        return False
    # 保守策略：仅在没有空格的串行情形下，从 item_name 自动拆分 model/unit。
    # 对于“DJI CP. Matrice 4T ...”这类空格分词文本，直接保留原始内容。
    if not has_spacing:
        if not unit:
            m_unit = re.search(r"(?P<unit>[A-Za-z]{1,4}|[\u4e00-\u9fff]{1,3})$", compact_item)
            if m_unit:
                unit_candidate = m_unit.group("unit")
                prefix = compact_item[: m_unit.start()].strip()
                if prefix and looks_like_unit_token(unit_candidate) and re.search(r"[A-Za-z0-9.\-*/()]$", prefix):
                    unit = unit_candidate
                    compact_item = prefix
        if not model:
            m_model = re.search(r"([A-Za-z0-9][A-Za-z0-9.\-*/]{1,})$", compact_item)
            if m_model:
                candidate = m_model.group(1).strip(".-*/")
                if len(candidate) >= 2 and re.search(r"[A-Za-z0-9]", candidate):
                    model = candidate
                    compact_item = compact_item[: m_model.start()].strip()
        if compact_item and not model:
            packed = re.search(r"(?P<name>.+?)(?P<model>[A-Za-z0-9][A-Za-z0-9.\-*/()（）]{2,})(?P<unit>[A-Za-z]{1,4}|[\u4e00-\u9fff]{1,3})?$", compact_item)
            if packed:
                name_candidate = packed.group("name").strip()
                model_candidate = packed.group("model").strip(".-*/")
                unit_candidate = (packed.group("unit") or "").strip()
                if name_candidate and model_candidate and re.search(r"[A-Za-z]", model_candidate):
                    compact_item = name_candidate
                    model = model_candidate
                    if unit_candidate and not unit and looks_like_unit_token(unit_candidate):
                        unit = unit_candidate
        item_name = compact_item
    item_name = re.sub(r"\s+", " ", item_name).strip()
    model = re.sub(r"\s+", " ", model).strip()
    unit = re.sub(r"\s+", " ", unit).strip()
    return {"item_name": item_name, "model": model, "unit": unit}
def extract_header_fields(text: str) -> Dict[str, str]:
    compact = compact_label(text)
    def clean_party_name(name: str) -> str:
        name = (name or "").strip()
        if not name:
            return ""
        name = re.sub(r"^名称[:：]?", "", name).strip()
        # 清除误识别到名称末尾的“购买方信息/销售方信息/买售方方信”等标签
        name = re.sub(
            r"[（(]\s*(?:购买方信息|销售方信息|买方信息|卖方信息|买售方方信)\s*$",
            "",
            name,
        ).strip()
        # 防止“销 名称:xxx”或“销售方名称:xxx”等串到同一字段
        name = re.split(r"(?:销\s*名称|销售方\s*名称)\s*[:：]", name, maxsplit=1)[0].strip()
        # 防止税号、地址等内容误拼到名称里
        name = re.split(
            r"(?:统一社会信用代码/?纳税人识别号|纳税人识别号|地址|开户行|电话|账号|购买方信息|销售方信息|买方信息|卖方信息|买售方方信|价税合计)",
            name,
            maxsplit=1,
        )[0].strip()
        return name
    def extract_party_fields(raw_text: str, party: str) -> Dict[str, str]:
        # 兼容“购买方信息/购 买 方 信 息/竖排换行”等布局
        party_map = {
            "buyer": ("购买方信息", "销售方信息"),
            "seller": ("销售方信息", "价税合计"),
        }
        begin, end = party_map[party]
        block = first_match(
            rf"{begin}(.*?)(?:{end}|$)",
            compact_label(raw_text),
            flags=re.S,
        )
        name = first_match(r"名称[:：]?([^\n]+?)(?:统一社会信用代码|纳税人识别号|地址|开户行|$)", block)
        tax_no = first_match(r"(?:统一社会信用代码/?纳税人识别号|纳税人识别号)[:：]?([0-9A-Z]{15,20})", block)
        return {"name": clean_party_name(name), "tax_no": tax_no}
    fields = {
        "invoice_number": first_match(r"发票号码\s*:\s*([0-9A-Za-z]+)", text),
        "invoice_date": first_match(r"开票日期\s*:\s*([0-9]{4}[-/年][0-9]{1,2}[-/月][0-9]{1,2}日?)", text),
        "buyer_name": "",
        "buyer_tax_no": "",
        "seller_name": "",
        "seller_tax_no": "",
        "total": "",
    }
    buyer_fields = extract_party_fields(text, "buyer")
    seller_fields = extract_party_fields(text, "seller")
    fields["buyer_name"] = buyer_fields["name"]
    fields["buyer_tax_no"] = buyer_fields["tax_no"]
    fields["seller_name"] = seller_fields["name"]
    fields["seller_tax_no"] = seller_fields["tax_no"]
    # 兜底：某些模板无法可靠切块时，名称/税号按出现顺序取第1个买方、第2个卖方
    if not fields["buyer_name"] or not fields["seller_name"]:
        names = re.findall(r"名称\s*[:：]\s*([^\n]+)", text)
        if len(names) >= 1 and not fields["buyer_name"]:
            fields["buyer_name"] = clean_party_name(names[0])
        if len(names) >= 2 and not fields["seller_name"]:
            fields["seller_name"] = clean_party_name(names[1])
    # 再兜底：压缩文本里按“名称:”顺序提取，避免同一行串到一起
    if not fields["buyer_name"] or not fields["seller_name"]:
        compact_names = re.findall(
            r"名称[:：](.*?)(?=(?:统一社会信用代码/?纳税人识别号|纳税人识别号|地址|开户行|销售方信息|价税合计|名称[:：]|$))",
            compact,
        )
        compact_names = [clean_party_name(n) for n in compact_names if clean_party_name(n)]
        if len(compact_names) >= 1 and not fields["buyer_name"]:
            fields["buyer_name"] = compact_names[0]
        if len(compact_names) >= 2 and not fields["seller_name"]:
            fields["seller_name"] = compact_names[1]
    fields["buyer_name"] = clean_party_name(fields["buyer_name"])
    fields["seller_name"] = clean_party_name(fields["seller_name"])
    if not fields["buyer_tax_no"] or not fields["seller_tax_no"]:
        tax_list = re.findall(r"(?:统一社会信用代码/?纳税人识别号|纳税人识别号)\s*[:：]\s*([0-9A-Z]{15,20})", compact)
        if len(tax_list) >= 1 and not fields["buyer_tax_no"]:
            fields["buyer_tax_no"] = tax_list[0].strip()
        if len(tax_list) >= 2 and not fields["seller_tax_no"]:
            fields["seller_tax_no"] = tax_list[1].strip()
    total = first_match(r"\(小写\)\s*[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)", text)
    fields["total"] = clean_money(total)
    return fields
def open_output_path(path: str):
    try:
        if os.name == "nt":
            os.startfile(path)
            return
        if os.name == "posix":
            import subprocess
            import sys
            if sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
    except Exception:
        # 打开失败不影响主流程
        pass
def _clean_cell(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()
def extract_items_from_tables(pdf_path: str) -> List[Dict[str, str]]:
    if pdfplumber is None:
        return []
    items: List[Dict[str, str]] = []
    header_alias = {
        "item_name": ["项目名称", "货物或应税劳务、服务名称", "服务名称", "产品名称"],
        "model": ["规格型号", "型号"],
        "unit": ["单位"],
        "quantity": ["数量"],
        "unit_price": ["单价"],
        "amount": ["金额", "不含税金额"],
        "tax_rate": ["税率", "征收率", "税率/征收率"],
        "tax_amount": ["税额"],
    }
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table:
                    continue
                rows = [[_clean_cell(cell) for cell in (row or [])] for row in table]
                header_idx = -1
                col_map: Dict[str, int] = {}
                for idx, row in enumerate(rows):
                    joined = " ".join(row)
                    if "项目名称" not in joined and "产品名称" not in joined:
                        continue
                    tmp_map: Dict[str, int] = {}
                    for col_idx, col_name in enumerate(row):
                        for key, aliases in header_alias.items():
                            if any(alias in col_name for alias in aliases):
                                tmp_map[key] = col_idx
                    if "item_name" in tmp_map and ("amount" in tmp_map or "tax_amount" in tmp_map):
                        header_idx = idx
                        col_map = tmp_map
                        break
                if header_idx < 0:
                    continue
                pending_item: Optional[Dict[str, str]] = None
                def merge_text_row(target: Dict[str, str], text_row: Dict[str, str]):
                    text_name = (text_row.get("item_name") or "").strip()
                    text_model = (text_row.get("model") or "").strip()
                    text_unit = (text_row.get("unit") or "").strip()
                    if text_name:
                        # 折行延续：若是纯中文描述，优先补到品名；含英文/数字更可能是规格型号。
                        if re.search(r"[A-Za-z0-9]", text_name):
                            if target.get("model"):
                                target["model"] = (target["model"] + " " + text_name).strip()
                            else:
                                target["model"] = text_name
                        else:
                            target["item_name"] = (target.get("item_name", "") + text_name).strip()
                    if text_model:
                        target["model"] = (target.get("model", "") + " " + text_model).strip()
                    if text_unit and not target.get("unit"):
                        target["unit"] = text_unit
                for row in rows[header_idx + 1 :]:
                    if not any(row):
                        continue
                    joined = "".join(row)
                    if any(stop in joined for stop in ["价税合计", "合计", "备注", "销售方信息", "购买方信息"]):
                        continue
                    def pick(key: str) -> str:
                        idx = col_map.get(key, -1)
                        return row[idx] if 0 <= idx < len(row) else ""
                    item_name = pick("item_name").lstrip("*")
                    amount = clean_money(pick("amount"))
                    tax_amount = clean_money(pick("tax_amount"))
                    quantity = pick("quantity")
                    unit_price = clean_money(pick("unit_price"))
                    tax_rate = pick("tax_rate")
                    model = pick("model")
                    unit = pick("unit")
                    has_text = bool(item_name or model or unit)
                    has_numbers = bool(quantity or unit_price or amount or tax_rate or tax_amount)
                    current = {
                        "item_name": item_name,
                        "model": model,
                        "unit": unit,
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "amount": amount,
                        "tax_rate": tax_rate,
                        "tax_amount": tax_amount,
                    }
                    if pending_item is not None:
                        if has_numbers and not item_name:
                            for key in ["quantity", "unit_price", "amount", "tax_rate", "tax_amount"]:
                                if current[key]:
                                    pending_item[key] = current[key]
                            if any(pending_item[k] for k in ["item_name", "amount", "tax_amount"]):
                                items.append(pending_item)
                            pending_item = None
                            continue
                        if has_numbers:
                            merged = dict(current)
                            merge_text_row(merged, pending_item)
                            items.append(merged)
                            pending_item = None
                            continue
                        if has_text and not has_numbers:
                            merge_text_row(pending_item, current)
                            continue
                        if any(pending_item[k] for k in ["item_name", "amount", "tax_amount"]):
                            items.append(pending_item)
                        pending_item = None
                    if not has_text and not has_numbers:
                        continue
                    if has_text and not has_numbers:
                        # 常见场景：上一行已包含数量金额，本行仅是折行规格型号；应回填上一条，避免拆成新行。
                        if items and any(items[-1].get(k) for k in ["quantity", "unit_price", "amount", "tax_rate", "tax_amount"]):
                            merge_text_row(items[-1], current)
                        else:
                            pending_item = current
                        continue
                    items.append(current)
                if pending_item is not None and any(pending_item[k] for k in ["item_name", "amount", "tax_amount"]):
                    items.append(pending_item)
    return items
def extract_items(text: str) -> List[Dict[str, str]]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    items = []
    def parse_item_line(line: str) -> Optional[Dict[str, str]]:
        def looks_like_unit_token(token: str) -> bool:
            token = (token or "").strip()
            if not token:
                return False
            if re.search(r"\d", token):
                return False
            if re.search(r"[*/_=+~`!@#$%^&?,;:'\"\\|]", token):
                return False
            if re.fullmatch(r"[A-Za-z]{1,4}", token):
                return True
            if re.fullmatch(r"[\u4e00-\u9fff]{1,3}", token):
                return True
            return False
        def looks_like_model_token(token: str) -> bool:
            token = (token or "").strip()
            if len(token) < 2:
                return False
            return bool(re.search(r"[A-Za-z0-9]", token))
        line = re.sub(r"\s+", " ", line).strip()
        if not line:
            return None
        # 从右向左提取金额列，避免产品名/规格型号/单位串行污染前半段。
        number_pat = r"[0-9]+(?:\.[0-9]+)?"
        tail = re.search(
            rf"(?P<quantity>{number_pat})\s+"
            rf"(?P<unit_price>{number_pat})\s+"
            rf"(?P<amount>{number_pat})\s+"
            rf"(?P<tax_rate>{number_pat}%?)\s+"
            rf"(?P<tax_amount>{number_pat})$",
            line,
        )
        if not tail:
            return None
        prefix = line[: tail.start()].strip().lstrip("*").strip()
        if not prefix:
            return None
        chunks = prefix.split()
        unit = ""
        model = ""
        item_name = prefix
        if len(chunks) >= 2:
            possible_unit = chunks[-1]
            possible_model = chunks[-2]
            if looks_like_unit_token(possible_unit):
                unit = possible_unit
                if looks_like_model_token(possible_model):
                    model = possible_model
                    item_name = " ".join(chunks[:-2]).strip()
                else:
                    item_name = " ".join(chunks[:-1]).strip()
            elif looks_like_model_token(possible_unit):
                model = possible_unit
                item_name = " ".join(chunks[:-1]).strip()
        if not item_name:
            item_name = prefix
        return {
            "item_name": item_name,
            "model": model,
            "unit": unit,
            "quantity": tail.group("quantity"),
            "unit_price": tail.group("unit_price"),
            "amount": tail.group("amount"),
            "tax_rate": tail.group("tax_rate"),
            "tax_amount": tail.group("tax_amount"),
        }
    # 先用跨行明细分段提取，兼容“项目名称换行 + 数字在下一行”的情况
    block = first_match(r"(?:项目名称|产品名称).*?(?=价税合计|合计|备注|$)", text, flags=re.S, group=0)
    if block:
        detail_lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        current_idx = -1
        for ln in detail_lines:
            ln_compact = compact_label(ln)
            if any(flag in ln_compact for flag in ["项目名称", "规格型号", "税率/征收率", "税额"]):
                continue
            if any(stop in ln_compact for stop in ["价税合计", "备注", "开票人"]) or ln_compact.startswith("合计"):
                break
            parsed = parse_item_line(ln)
            if parsed:
                items.append(parsed)
                current_idx = len(items) - 1
                continue
            if current_idx < 0:
                continue
            extra = re.sub(r"\s+", " ", ln).strip().lstrip("*")
            if not extra:
                continue
            # 折行补全：优先补到规格型号，若明显是中文品名描述则补到品名
            if re.search(r"[A-Za-z0-9]", extra) or any(ch in extra for ch in ["(", ")", "（", "）", ".", "-"]):
                target = "model"
            else:
                target = "item_name"
            items[current_idx][target] = (items[current_idx][target] + " " + extra).strip()
    if items:
        return items
    for line in lines:
        if "*" not in line:
            continue
        if "项目名称" in line or "规格型号" in line:
            continue
        line_norm = re.sub(r"\s+", " ", line)
        parsed = parse_item_line(line_norm)
        if parsed:
            items.append(parsed)
            continue
        fallback = re.match(
            r"^(\*[^\s]+(?:\s+[^\s]+)*)\s+.*?([0-9]+(?:\.[0-9]+)?)\s+([0-9]+(?:\.[0-9]+)?)$",
            line_norm,
        )
        if fallback:
            items.append(
                {
                    "item_name": fallback.group(1).lstrip("*"),
                    "model": "",
                    "unit": "",
                    "quantity": "",
                    "unit_price": "",
                    "amount": fallback.group(2),
                    "tax_rate": "",
                    "tax_amount": fallback.group(3),
                }
            )
        # 纯文本明细（无数量/金额）兜底：如“交通运输设备*游艇拖车RIB550QR台”
        pure_text = re.match(r"^(\*?\S.*)$", line_norm)
        if pure_text and not re.search(r"\d+\s+\d+", line_norm):
            raw = pure_text.group(1).lstrip("*").strip()
            if raw and any(ch in raw for ch in ["*", "设备", "产品", "装置"]):
                items.append(
                    {
                        "item_name": raw,
                        "model": "",
                        "unit": "",
                        "quantity": "",
                        "unit_price": "",
                        "amount": "",
                        "tax_rate": "",
                        "tax_amount": "",
                    }
                )
    return items
def parse_invoice(pdf_path: str) -> List[InvoiceRow]:
    structured = extract_structured_invoice(pdf_path)
    if structured:
        header = structured["header"]
        items = structured["items"]
    else:
        text = extract_text_from_pdf(pdf_path)
        header = extract_header_fields(text)
        items = extract_items_from_tables(pdf_path)
        if not items:
            items = extract_items(text)
    if not items:
        items = [{
            "item_name": "",
            "model": "",
            "unit": "",
            "quantity": "",
            "unit_price": "",
            "amount": "",
            "tax_rate": "",
            "tax_amount": "",
        }]
    rows = []
    for item in items:
        normalized_item = split_name_model_unit(item.get("item_name", ""), item.get("model", ""), item.get("unit", ""))
        rows.append(
            InvoiceRow(
                source_file=os.path.basename(pdf_path),
                invoice_number=header.get("invoice_number", ""),
                invoice_date=header.get("invoice_date", ""),
                buyer_name=header.get("buyer_name", ""),
                buyer_tax_no=header.get("buyer_tax_no", ""),
                seller_name=header.get("seller_name", ""),
                seller_tax_no=header.get("seller_tax_no", ""),
                item_name=normalized_item["item_name"],
                model=normalized_item["model"],
                unit=normalized_item["unit"],
                quantity=item.get("quantity", ""),
                unit_price=item.get("unit_price", ""),
                amount=item.get("amount", ""),
                tax_rate=item.get("tax_rate", ""),
                tax_amount=item.get("tax_amount", ""),
                total=header.get("total", ""),
            )
        )
    return rows
def save_to_csv(rows: List[InvoiceRow], out_path: str):
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_export_dict())
def beautify_excel(path: str):
    if load_workbook is None:
        return
    wb = load_workbook(path)
    ws = wb.active
    header_fill = PatternFill(fill_type="solid", fgColor="2F75B5")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 4, 12), 45)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)
def save_to_xlsx(rows: List[InvoiceRow], out_path: str):
    if pd is None:
        raise RuntimeError("缺少依赖: pandas/openpyxl，请先安装 (pip install pandas openpyxl)")
    df = pd.DataFrame([r.to_export_dict() for r in rows], columns=EXPORT_COLUMNS)
    df.to_excel(out_path, index=False)
    beautify_excel(out_path)
class InvoiceApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("发票PDF批量整理助手")
        self.root.geometry("980x680")
        self.pdf_files: List[str] = []
        self.mode_var = tk.StringVar(value="merge")
        self.format_var = tk.StringVar(value="xlsx")
        self._build_ui()
    def _build_ui(self):
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)
        top = ttk.Frame(main)
        top.pack(fill="x")
        ttk.Label(top, text="1) 选择并调整要处理的PDF文件顺序", font=("Microsoft YaHei", 11, "bold")).pack(anchor="w")
        list_wrap = ttk.Frame(main)
        list_wrap.pack(fill="both", expand=True, pady=10)
        self.listbox = tk.Listbox(list_wrap, selectmode=tk.EXTENDED, font=("Microsoft YaHei", 10))
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(list_wrap, orient="vertical", command=self.listbox.yview)
        scrollbar.pack(side="left", fill="y")
        self.listbox.config(yscrollcommand=scrollbar.set)
        btns = ttk.Frame(list_wrap)
        btns.pack(side="left", padx=10, fill="y")
        ttk.Button(btns, text="添加PDF", command=self.add_files).pack(fill="x", pady=2)
        ttk.Button(btns, text="移除选中", command=self.remove_selected).pack(fill="x", pady=2)
        ttk.Button(btns, text="上移", command=self.move_up).pack(fill="x", pady=2)
        ttk.Button(btns, text="下移", command=self.move_down).pack(fill="x", pady=2)
        ttk.Button(btns, text="清空", command=self.clear_files).pack(fill="x", pady=2)
        options = ttk.LabelFrame(main, text="2) 导出设置", padding=10)
        options.pack(fill="x", pady=8)
        ttk.Radiobutton(options, text="合并整理（一个总表）", variable=self.mode_var, value="merge").grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(options, text="分开整理（每个PDF单独表）", variable=self.mode_var, value="split").grid(row=0, column=1, sticky="w", padx=20)
        ttk.Label(options, text="输出格式:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Radiobutton(options, text="XLSX（推荐，美观）", variable=self.format_var, value="xlsx").grid(row=1, column=1, sticky="w", pady=(8, 0))
        ttk.Radiobutton(options, text="CSV（UTF-8带BOM）", variable=self.format_var, value="csv").grid(row=1, column=2, sticky="w", pady=(8, 0), padx=15)
        progress_wrap = ttk.LabelFrame(main, text="3) 处理进度", padding=10)
        progress_wrap.pack(fill="x", pady=8)
        self.progress = ttk.Progressbar(progress_wrap, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(fill="x")
        self.status_var = tk.StringVar(value="等待开始")
        ttk.Label(progress_wrap, textvariable=self.status_var).pack(anchor="w", pady=(6, 0))
        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=10)
        ttk.Button(bottom, text="开始整理并导出", command=self.start_process).pack(side="left")
        ttk.Button(bottom, text="退出", command=self.root.destroy).pack(side="right")
        tip = (
            "说明: 若部分PDF是扫描件（图片）而非文本，需先OCR后再识别。"
            "如果提取结果不完整，请把样例PDF（或脱敏截图）提供给开发者调整规则。"
        )
        ttk.Label(main, text=tip, foreground="#555555").pack(anchor="w", pady=(6, 0))
    def refresh_listbox(self):
        self.listbox.delete(0, tk.END)
        for p in self.pdf_files:
            self.listbox.insert(tk.END, p)
    def add_files(self):
        paths = filedialog.askopenfilenames(title="选择PDF文件", filetypes=[("PDF 文件", "*.pdf")])
        if not paths:
            return
        for p in paths:
            if p not in self.pdf_files:
                self.pdf_files.append(p)
        self.refresh_listbox()
    def remove_selected(self):
        selected = list(self.listbox.curselection())
        for idx in reversed(selected):
            del self.pdf_files[idx]
        self.refresh_listbox()
    def move_up(self):
        selected = list(self.listbox.curselection())
        if not selected:
            return
        for idx in selected:
            if idx == 0:
                continue
            self.pdf_files[idx - 1], self.pdf_files[idx] = self.pdf_files[idx], self.pdf_files[idx - 1]
        self.refresh_listbox()
        for idx in [max(0, i - 1) for i in selected]:
            self.listbox.selection_set(idx)
    def move_down(self):
        selected = list(self.listbox.curselection())
        if not selected:
            return
        for idx in reversed(selected):
            if idx >= len(self.pdf_files) - 1:
                continue
            self.pdf_files[idx + 1], self.pdf_files[idx] = self.pdf_files[idx], self.pdf_files[idx + 1]
        self.refresh_listbox()
        for idx in [min(len(self.pdf_files) - 1, i + 1) for i in selected]:
            self.listbox.selection_set(idx)
    def clear_files(self):
        self.pdf_files.clear()
        self.refresh_listbox()
    def start_process(self):
        if not self.pdf_files:
            messagebox.showwarning("提示", "请先添加至少一个PDF文件")
            return
        if pdfplumber is None:
            messagebox.showerror("缺少依赖", "请先安装 pdfplumber:\npip install pdfplumber")
            return
        if self.format_var.get() == "xlsx" and pd is None:
            messagebox.showerror("缺少依赖", "导出XLSX需安装 pandas/openpyxl:\npip install pandas openpyxl")
            return
        out_dir = filedialog.askdirectory(title="选择导出目录")
        if not out_dir:
            return
        threading.Thread(target=self._process_worker, args=(out_dir,), daemon=True).start()
    def _process_worker(self, out_dir: str):
        total = len(self.pdf_files)
        self.progress.configure(maximum=total, value=0)
        all_rows: List[InvoiceRow] = []
        try:
            for idx, pdf_path in enumerate(self.pdf_files, start=1):
                self.status_var.set(f"正在处理 {idx}/{total}: {os.path.basename(pdf_path)}")
                rows = parse_invoice(pdf_path)
                all_rows.extend(rows)
                if self.mode_var.get() == "split":
                    base = os.path.splitext(os.path.basename(pdf_path))[0]
                    ext = self.format_var.get()
                    out_path = os.path.join(out_dir, f"{base}_整理.{ext}")
                    if ext == "csv":
                        save_to_csv(rows, out_path)
                    else:
                        save_to_xlsx(rows, out_path)
                self.progress.configure(value=idx)
            if self.mode_var.get() == "merge":
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                ext = self.format_var.get()
                out_path = os.path.join(out_dir, f"发票汇总_{stamp}.{ext}")
                if ext == "csv":
                    save_to_csv(all_rows, out_path)
                else:
                    save_to_xlsx(all_rows, out_path)
            self.status_var.set("处理完成")
            messagebox.showinfo("完成", f"导出完成，共处理 {total} 个PDF。")
            if self.mode_var.get() == "merge":
                open_output_path(out_path)
            elif total == 1 and self.mode_var.get() == "split":
                open_output_path(os.path.join(out_dir, f"{os.path.splitext(os.path.basename(self.pdf_files[0]))[0]}_整理.{self.format_var.get()}"))
        except Exception as e:
            self.status_var.set("处理失败")
            messagebox.showerror("错误", f"处理过程中出现问题:\n{e}")
def main():
    root = tk.Tk()
    try:
        style = ttk.Style(root)
        if "vista" in style.theme_names():
            style.theme_use("vista")
        elif "clam" in style.theme_names():
            style.theme_use("clam")
    except Exception:
        pass
    app = InvoiceApp(root)
    root.mainloop()
if __name__ == "__main__":
    main()
